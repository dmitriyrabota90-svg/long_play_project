from __future__ import annotations

import os
import subprocess
import sys
from datetime import date, datetime, timezone
from decimal import Decimal
from io import BytesIO
from pathlib import Path

import httpx
import pytest
from openpyxl import Workbook, load_workbook
from sqlalchemy import create_engine, func, select
from sqlalchemy.orm import sessionmaker

from app.collectors.benchmarks.world_bank_pink_sheet import (
    WorldBankPinkSheetCollector,
    WorldBankPinkSheetParseError,
    build_commodity_benchmark_source_record_hash,
    parse_world_bank_benchmark,
)
from app.collectors.benchmarks.world_bank_pink_sheet_config import (
    REQUEST_TIMEOUT,
    SOURCE_FILE_URL,
    WORLD_BANK_BENCHMARKS,
    WORLD_BANK_BENCHMARKS_BY_CODE,
)
from app.db.models import Base, CollectorRun, CommodityBenchmark, DataQualityCheck, RawResponse
from app.db.seed import seed_database
from app.storage.raw_store import RawStore
from scripts.run_collector import build_parser


def _pink_sheet_xlsx(*, soybean_oil_value: str | float = "1000.50") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Monthly Prices"
    ws["A1"] = "World Bank Commodity Price Data (The Pink Sheet)"
    ws["A4"] = "Updated on June 02, 2026"
    ws["B5"] = "Soybean oil"
    ws["C5"] = "Soybeans"
    ws["D5"] = "Palm oil"
    ws["E5"] = "Maize"
    ws["F5"] = "Wheat, US SRW"
    for col in "BCDEF":
        ws[f"{col}6"] = "($/mt)"
    ws["A7"] = "2024M01"
    ws["B7"] = soybean_oil_value
    ws["C7"] = "500.25"
    ws["D7"] = "900"
    ws["E7"] = "220"
    ws["F7"] = "240"
    ws["A8"] = "2024M02"
    ws["B8"] = "…"
    ws["C8"] = "bad"
    ws["D8"] = None
    ws["E8"] = "225"
    ws["F8"] = "245"
    ws["A9"] = "2024M03"
    ws["B9"] = "1010.75"
    ws["C9"] = "510.25"
    ws["D9"] = "910"
    ws["E9"] = "230"
    ws["F9"] = "250"

    idx = wb.create_sheet("Monthly Indices")
    idx["A1"] = "World Bank Commodity Price Data (The Pink Sheet)"
    idx["N7"] = "Fertilizers **"
    idx["A10"] = "2024M01"
    idx["N10"] = "120.5"
    idx["A11"] = "2024M02"
    idx["N11"] = "121.5"

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


class FakeHttpClient:
    def __init__(self, payload: bytes | httpx.Response | None = None) -> None:
        self.payload = payload if payload is not None else _pink_sheet_xlsx()
        self.calls: list[tuple[str, dict[str, str], float]] = []

    def get(self, url: str, *, headers: dict[str, str], timeout: float) -> httpx.Response:
        self.calls.append((url, headers, timeout))
        if isinstance(self.payload, httpx.Response):
            return self.payload
        return httpx.Response(
            200,
            content=self.payload,
            headers={"content-type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
            request=httpx.Request("GET", url),
        )


def test_benchmark_config_contains_expected_codes() -> None:
    assert {
        "world_bank_soybean_oil",
        "world_bank_soybeans",
        "world_bank_palm_oil",
        "world_bank_maize",
        "world_bank_wheat",
        "world_bank_fertilizer_index",
    } <= set(WORLD_BANK_BENCHMARKS_BY_CODE)
    assert SOURCE_FILE_URL.endswith("CMO-Historical-Data-Monthly.xlsx")
    assert WORLD_BANK_BENCHMARKS_BY_CODE["world_bank_soybean_oil"].unit == "USD/metric_ton"
    assert WORLD_BANK_BENCHMARKS_BY_CODE["world_bank_fertilizer_index"].unit == "index_2010_100"


def test_parser_parses_monthly_prices_and_filters_dates() -> None:
    workbook = load_workbook(BytesIO(_pink_sheet_xlsx()), read_only=True, data_only=True)
    benchmark = WORLD_BANK_BENCHMARKS_BY_CODE["world_bank_soybean_oil"]

    rows, source_column = parse_world_bank_benchmark(
        workbook,
        benchmark=benchmark,
        from_date=date(2024, 1, 1),
        to_date=date(2024, 2, 1),
    )

    assert source_column == "Soybean oil"
    assert len(rows) == 1
    assert rows[0].period_start == date(2024, 1, 1)
    assert rows[0].period_end == date(2024, 1, 31)
    assert rows[0].observed_at == datetime(2024, 1, 31, tzinfo=timezone.utc)
    assert rows[0].value == Decimal("1000.50")


def test_parser_parses_index_sheet() -> None:
    workbook = load_workbook(BytesIO(_pink_sheet_xlsx()), read_only=True, data_only=True)
    benchmark = WORLD_BANK_BENCHMARKS_BY_CODE["world_bank_fertilizer_index"]

    rows, source_column = parse_world_bank_benchmark(workbook, benchmark=benchmark, to_date=date(2024, 1, 1))

    assert source_column == "Fertilizers **"
    assert len(rows) == 1
    assert rows[0].value == Decimal("120.5")


def test_parser_reports_unknown_format() -> None:
    workbook = load_workbook(BytesIO(_pink_sheet_xlsx()), read_only=True, data_only=True)
    benchmark = WORLD_BANK_BENCHMARKS_BY_CODE["world_bank_soybean_oil"]
    missing_column = benchmark.__class__(
        **{**benchmark.__dict__, "source_column_candidates": ("No such column",)}
    )

    with pytest.raises(WorldBankPinkSheetParseError, match="source column not found"):
        parse_world_bank_benchmark(workbook, benchmark=missing_column)


def test_collector_inserts_benchmarks_with_raw_and_quality_checks(tmp_path: Path) -> None:
    session, cleanup = build_seeded_session()
    client = FakeHttpClient(_pink_sheet_xlsx())
    try:
        result = WorldBankPinkSheetCollector(
            benchmarks=(WORLD_BANK_BENCHMARKS_BY_CODE["world_bank_soybean_oil"],),
            http_client=client,
            raw_store=RawStore(base_dir=tmp_path / "raw"),
            clock=lambda: datetime(2026, 6, 9, 12, tzinfo=timezone.utc),
        ).run(
            session,
            benchmark_code="world_bank_soybean_oil",
            from_date=date(2024, 1, 1),
            to_date=date(2024, 3, 1),
        )
        rows = session.scalars(select(CommodityBenchmark).order_by(CommodityBenchmark.period_start)).all()
        raw_response = session.scalar(select(RawResponse))
        run = session.scalar(select(CollectorRun))
        check_names = set(session.scalars(select(DataQualityCheck.check_name)).all())
    finally:
        cleanup()

    assert result.status == "success"
    assert result.benchmarks_requested == 1
    assert result.benchmarks_success == 1
    assert result.records_found == 2
    assert result.records_written == 2
    assert result.skipped_existing == 0
    assert result.conflicts_count == 0
    assert rows[0].benchmark_code == "world_bank_soybean_oil"
    assert rows[0].value == Decimal("1000.50000000")
    assert rows[0].currency == "USD"
    assert rows[0].unit == "USD/metric_ton"
    assert rows[0].normalized_value == rows[0].value
    assert rows[0].metadata_json["source_column"] == "Soybean oil"
    assert raw_response is not None
    assert run is not None
    assert all(row.raw_response_id == raw_response.id for row in rows)
    assert all(row.collector_run_id == run.id for row in rows)
    assert client.calls[0][0] == SOURCE_FILE_URL
    assert client.calls[0][2] == REQUEST_TIMEOUT
    assert "commodity_benchmark_raw_response_non_empty" in check_names
    assert "commodity_benchmark_mapping_found" in check_names
    assert "commodity_benchmark_value_positive" in check_names


def test_duplicate_retry_skips_existing_rows(tmp_path: Path) -> None:
    session, cleanup = build_seeded_session()
    try:
        collector = WorldBankPinkSheetCollector(
            benchmarks=(WORLD_BANK_BENCHMARKS_BY_CODE["world_bank_soybean_oil"],),
            http_client=FakeHttpClient(_pink_sheet_xlsx()),
            raw_store=RawStore(base_dir=tmp_path / "raw"),
            clock=lambda: datetime(2026, 6, 9, 12, tzinfo=timezone.utc),
        )
        first = collector.run(session, from_date=date(2024, 1, 1), to_date=date(2024, 1, 1))
        second = collector.run(session, from_date=date(2024, 1, 1), to_date=date(2024, 1, 1))
        rows_count = session.scalar(select(func.count()).select_from(CommodityBenchmark))
    finally:
        cleanup()

    assert first.records_written == 1
    assert second.records_written == 0
    assert second.skipped_existing == 1
    assert second.conflicts_count == 0
    assert rows_count == 1


def test_changed_existing_value_creates_conflict_warning_without_overwrite(tmp_path: Path) -> None:
    session, cleanup = build_seeded_session()
    try:
        WorldBankPinkSheetCollector(
            benchmarks=(WORLD_BANK_BENCHMARKS_BY_CODE["world_bank_soybean_oil"],),
            http_client=FakeHttpClient(_pink_sheet_xlsx(soybean_oil_value="1000.50")),
            raw_store=RawStore(base_dir=tmp_path / "raw"),
            clock=lambda: datetime(2026, 6, 9, 12, tzinfo=timezone.utc),
        ).run(session, from_date=date(2024, 1, 1), to_date=date(2024, 1, 1))
        second = WorldBankPinkSheetCollector(
            benchmarks=(WORLD_BANK_BENCHMARKS_BY_CODE["world_bank_soybean_oil"],),
            http_client=FakeHttpClient(_pink_sheet_xlsx(soybean_oil_value="1001.00")),
            raw_store=RawStore(base_dir=tmp_path / "raw"),
            clock=lambda: datetime(2026, 6, 9, 12, 5, tzinfo=timezone.utc),
        ).run(session, from_date=date(2024, 1, 1), to_date=date(2024, 1, 1))
        row = session.scalar(select(CommodityBenchmark))
        warning = session.scalar(
            select(DataQualityCheck).where(
                DataQualityCheck.check_name == "commodity_benchmark_existing_record_hash_changed"
            )
        )
    finally:
        cleanup()

    assert second.status == "partial_success"
    assert second.records_written == 0
    assert second.conflicts_count == 1
    assert row.value == Decimal("1000.50000000")
    assert warning is not None
    assert warning.severity == "warning"
    assert warning.details_json["incoming_value"] == "1001"
    assert warning.details_json["policy_decision"] == "rejected_hash_conflict_no_overwrite"


def test_one_bad_benchmark_does_not_fail_whole_run(tmp_path: Path) -> None:
    session, cleanup = build_seeded_session()
    good = WORLD_BANK_BENCHMARKS_BY_CODE["world_bank_soybean_oil"]
    bad = good.__class__(**{**good.__dict__, "benchmark_code": "bad_benchmark", "source_column_candidates": ("Nope",)})
    try:
        result = WorldBankPinkSheetCollector(
            benchmarks=(good, bad),
            http_client=FakeHttpClient(_pink_sheet_xlsx()),
            raw_store=RawStore(base_dir=tmp_path / "raw"),
            clock=lambda: datetime(2026, 6, 9, 12, tzinfo=timezone.utc),
        ).run(session, from_date=date(2024, 1, 1), to_date=date(2024, 1, 1))
        rows_count = session.scalar(select(func.count()).select_from(CommodityBenchmark))
    finally:
        cleanup()

    assert result.status == "partial_success"
    assert result.benchmarks_success == 1
    assert result.benchmarks_failed == 1
    assert result.errors_count == 1
    assert result.records_written == 1
    assert rows_count == 1


def test_unknown_benchmark_rejected() -> None:
    collector = WorldBankPinkSheetCollector()

    with pytest.raises(ValueError, match="unknown World Bank Pink Sheet benchmark"):
        collector.run(benchmark_code="not_real")


def test_source_record_hash_uses_business_key_and_value() -> None:
    first = build_commodity_benchmark_source_record_hash(
        source_code="world_bank_pink_sheet",
        benchmark_code="world_bank_soybean_oil",
        frequency="monthly",
        period_start=date(2024, 1, 1),
        period_end=date(2024, 1, 31),
        value=Decimal("1000.50"),
        currency="USD",
        unit="USD/metric_ton",
    )
    second = build_commodity_benchmark_source_record_hash(
        source_code="world_bank_pink_sheet",
        benchmark_code="world_bank_soybean_oil",
        frequency="monthly",
        period_start=date(2024, 1, 1),
        period_end=date(2024, 1, 31),
        value=Decimal("1001.00"),
        currency="USD",
        unit="USD/metric_ton",
    )

    assert first != second


def test_cli_argument_parsing_supports_world_bank_pink_sheet() -> None:
    parser = build_parser()

    args = parser.parse_args(
        [
            "world_bank_pink_sheet",
            "--benchmark",
            "world_bank_soybean_oil",
            "--from-date",
            "2024-01-01",
            "--to-date",
            "2026-06-01",
        ]
    )

    assert args.collector_name == "world_bank_pink_sheet"
    assert args.benchmark == "world_bank_soybean_oil"
    assert args.from_date == "2024-01-01"
    assert args.to_date == "2026-06-01"


def test_cli_rejects_unknown_world_bank_benchmark_without_db_access(tmp_path: Path) -> None:
    result = subprocess.run(
        [
            sys.executable,
            "scripts/run_collector.py",
            "world_bank_pink_sheet",
            "--benchmark",
            "BAD_BENCHMARK",
        ],
        check=False,
        capture_output=True,
        text=True,
        env={**os.environ, "LOG_DIR": str(tmp_path / "logs")},
    )

    assert result.returncode != 0
    assert "unknown World Bank Pink Sheet benchmark: BAD_BENCHMARK" in result.stderr


def build_seeded_session():
    engine = create_engine("sqlite+pysqlite:///:memory:", future=True)
    Base.metadata.create_all(engine)
    SessionLocal = sessionmaker(bind=engine, future=True)
    session = SessionLocal()
    seed_database(session)

    def cleanup() -> None:
        session.close()
        engine.dispose()

    return session, cleanup
