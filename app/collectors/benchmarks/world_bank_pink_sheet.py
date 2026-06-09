from __future__ import annotations

import logging
from calendar import monthrange
from dataclasses import dataclass, field
from datetime import date, datetime, time, timezone
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any, Protocol

import httpx
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.collectors.base import BaseCollector, CollectorResult
from app.collectors.benchmarks.world_bank_pink_sheet_config import (
    COLLECTOR_NAME,
    HEADERS,
    MAX_RESPONSE_BYTES,
    PARSER_VERSION,
    REQUEST_TIMEOUT,
    SOURCE_CODE,
    SOURCE_DOCUMENT_URL,
    SOURCE_FILE_URL,
    WORLD_BANK_BENCHMARKS,
    WORLD_BANK_BENCHMARKS_BY_CODE,
    WorldBankBenchmark,
)
from app.db.models import CollectorRun, CommodityBenchmark, DataQualityCheck, RawResponse, Source
from app.db.session import session_scope
from app.storage.hashes import stable_json_hash
from app.storage.raw_store import RawStore

logger = logging.getLogger(__name__)

MISSING_VALUE_MARKERS = {"", ".", "..", "...", "…", "nan", "NaN", "NA", "N/A", "null", "NULL"}


class WorldBankPinkSheetParseError(ValueError):
    """Raised when the World Bank Pink Sheet workbook cannot be parsed safely."""


class HttpClient(Protocol):
    def get(self, url: str, *, headers: dict[str, str], timeout: float) -> httpx.Response:
        ...


@dataclass(frozen=True)
class ParsedCommodityBenchmark:
    benchmark: WorldBankBenchmark
    period_start: date
    period_end: date
    observed_at: datetime
    value: Decimal
    source_sheet: str
    source_column: str
    raw: dict[str, Any]


@dataclass(frozen=True)
class BenchmarkResponseCheck:
    check_name: str
    status: str
    severity: str
    details: dict[str, Any]


@dataclass(frozen=True)
class WorldBankPinkSheetCollectorResult(CollectorResult):
    benchmarks_requested: int = 0
    benchmarks_success: int = 0
    benchmarks_failed: int = 0
    skipped_existing: int = 0
    conflicts_count: int = 0
    benchmarks: tuple[str, ...] = ()
    from_date: str | None = None
    to_date: str | None = None


@dataclass
class _BenchmarkStats:
    benchmark_code: str
    records_found: int = 0
    records_written: int = 0
    skipped_existing: int = 0
    conflicts_count: int = 0
    errors: list[str] = field(default_factory=list)


class WorldBankPinkSheetCollector(BaseCollector):
    source_code = SOURCE_CODE
    collector_name = COLLECTOR_NAME
    parser_version = PARSER_VERSION

    def __init__(
        self,
        *,
        benchmarks: tuple[WorldBankBenchmark, ...] = WORLD_BANK_BENCHMARKS,
        http_client: HttpClient | None = None,
        raw_store: RawStore | None = None,
        timeout: float = REQUEST_TIMEOUT,
        clock: Any | None = None,
    ) -> None:
        super().__init__()
        self.benchmarks = benchmarks
        self.http_client = http_client or httpx.Client(follow_redirects=True)
        self.raw_store = raw_store or RawStore()
        self.timeout = timeout
        self.clock = clock or (lambda: datetime.now(timezone.utc))

    def run(
        self,
        session: Session | None = None,
        *,
        benchmark_code: str | None = None,
        from_date: date | None = None,
        to_date: date | None = None,
        run_type: str = "manual",
    ) -> WorldBankPinkSheetCollectorResult:
        run_type = _normalize_run_type(run_type)
        if run_type != "manual":
            raise ValueError("world_bank_pink_sheet is manual-only in Phase 6.2D")
        if from_date and to_date and from_date > to_date:
            raise ValueError("from_date must be on or before to_date")
        selected = _select_benchmarks(self.benchmarks, benchmark_code)
        if session is None:
            with session_scope() as scoped_session:
                return self._run_with_session(
                    scoped_session,
                    selected_benchmarks=selected,
                    from_date=from_date,
                    to_date=to_date,
                    run_type=run_type,
                )
        return self._run_with_session(
            session,
            selected_benchmarks=selected,
            from_date=from_date,
            to_date=to_date,
            run_type=run_type,
        )

    def _run_with_session(
        self,
        session: Session,
        *,
        selected_benchmarks: tuple[WorldBankBenchmark, ...],
        from_date: date | None,
        to_date: date | None,
        run_type: str,
    ) -> WorldBankPinkSheetCollectorResult:
        started_at = _to_utc(self.clock())
        source = session.scalar(select(Source).where(Source.code == self.source_code))
        collector_run = CollectorRun(
            source_id=source.id if source else None,
            collector_name=self.collector_name,
            started_at=started_at,
            status="running",
            run_type=run_type,
            collection_slot=None,
        )
        session.add(collector_run)
        session.flush()

        if source is None:
            message = f"source mapping not found: {self.source_code}"
            _write_quality_check(
                session,
                source_id=None,
                table_name="sources",
                check_name="commodity_benchmark_source_mapping_found",
                status="fail",
                severity="error",
                details={"source_code": self.source_code},
                checked_at=started_at,
            )
            collector_run.status = "error"
            collector_run.finished_at = _to_utc(self.clock())
            collector_run.error_message = message
            session.flush()
            return _result_from_run(
                self,
                collector_run,
                started_at,
                selected_benchmarks=selected_benchmarks,
                from_date=from_date,
                to_date=to_date,
                benchmarks_success=0,
                benchmarks_failed=len(selected_benchmarks),
                skipped_existing=0,
                conflicts_count=0,
                raw_response_ids=[],
                errors_count=1,
            )

        _write_quality_check(
            session,
            source_id=source.id,
            table_name="sources",
            check_name="commodity_benchmark_source_mapping_found",
            status="pass",
            severity="info",
            details={"source_code": self.source_code, "source_id": source.id},
            checked_at=started_at,
        )

        raw_response_ids: list[int] = []
        all_errors: list[str] = []
        records_found = 0
        records_written = 0
        skipped_existing = 0
        conflicts_count = 0
        benchmarks_success = 0
        benchmarks_failed = 0

        _write_quality_check(
            session,
            source_id=source.id,
            table_name="commodity_benchmarks",
            check_name="commodity_benchmark_date_range_valid",
            status="pass",
            severity="info",
            details={
                "from_date": from_date.isoformat() if from_date else None,
                "to_date": to_date.isoformat() if to_date else None,
                "filter_semantics": "period_start",
            },
            checked_at=started_at,
        )

        try:
            response = self.http_client.get(SOURCE_FILE_URL, headers=HEADERS, timeout=self.timeout)
            fetched_at = _to_utc(self.clock())
        except httpx.TimeoutException as exc:
            all_errors.append(f"World Bank Pink Sheet request timeout: {exc}")
            return self._finish_error_run(
                collector_run,
                started_at,
                selected_benchmarks=selected_benchmarks,
                from_date=from_date,
                to_date=to_date,
                errors=all_errors,
            )
        except httpx.RequestError as exc:
            all_errors.append(f"World Bank Pink Sheet connection error: {exc}")
            return self._finish_error_run(
                collector_run,
                started_at,
                selected_benchmarks=selected_benchmarks,
                from_date=from_date,
                to_date=to_date,
                errors=all_errors,
            )

        raw_response = self._save_raw_response(session, source, collector_run, response, fetched_at)
        raw_response_ids.append(raw_response.id)
        response_checks = assess_world_bank_response(
            payload=response.content,
            content_type=response.headers.get("content-type"),
            status_code=response.status_code,
        )
        _write_response_checks(
            session,
            source_id=source.id,
            raw_response_id=raw_response.id,
            checks=response_checks,
            checked_at=fetched_at,
        )
        if response.status_code != 200:
            all_errors.append(f"World Bank Pink Sheet HTTP status {response.status_code}")
        if not response.content:
            all_errors.append("World Bank Pink Sheet response is empty")
        if len(response.content) > MAX_RESPONSE_BYTES:
            all_errors.append(f"World Bank Pink Sheet response exceeds {MAX_RESPONSE_BYTES} bytes")
        if all_errors:
            return self._finish_error_run(
                collector_run,
                started_at,
                selected_benchmarks=selected_benchmarks,
                from_date=from_date,
                to_date=to_date,
                errors=all_errors,
                raw_response_ids=raw_response_ids,
            )

        try:
            workbook = load_workbook(BytesIO(response.content), read_only=True, data_only=True)
        except Exception as exc:  # pragma: no cover - openpyxl error types vary by version
            all_errors.append(f"World Bank Pink Sheet workbook parse error: {exc}")
            return self._finish_error_run(
                collector_run,
                started_at,
                selected_benchmarks=selected_benchmarks,
                from_date=from_date,
                to_date=to_date,
                errors=all_errors,
                raw_response_ids=raw_response_ids,
            )

        for benchmark in selected_benchmarks:
            stats = self._collect_benchmark(
                session,
                source,
                collector_run,
                raw_response,
                workbook,
                benchmark,
                from_date=from_date,
                to_date=to_date,
                fetched_at=fetched_at,
            )
            records_found += stats.records_found
            records_written += stats.records_written
            skipped_existing += stats.skipped_existing
            conflicts_count += stats.conflicts_count
            all_errors.extend(stats.errors)
            if stats.errors:
                benchmarks_failed += 1
            else:
                benchmarks_success += 1

        collector_run.finished_at = _to_utc(self.clock())
        collector_run.records_found = records_found
        collector_run.records_written = records_written
        if not all_errors and conflicts_count == 0 and benchmarks_success == len(selected_benchmarks):
            collector_run.status = "success"
        elif records_found > 0 or benchmarks_success > 0:
            collector_run.status = "partial_success"
        else:
            collector_run.status = "error"
        collector_run.error_message = "\n".join(all_errors) if all_errors else None
        session.flush()

        return WorldBankPinkSheetCollectorResult(
            run_id=collector_run.id,
            source_code=self.source_code,
            collector_name=self.collector_name,
            parser_version=self.parser_version,
            started_at=started_at,
            finished_at=collector_run.finished_at,
            status=collector_run.status,
            records_found=records_found,
            records_written=records_written,
            raw_response_ids=raw_response_ids,
            error_message=collector_run.error_message,
            errors_count=len(all_errors),
            benchmarks_requested=len(selected_benchmarks),
            benchmarks_success=benchmarks_success,
            benchmarks_failed=benchmarks_failed,
            skipped_existing=skipped_existing,
            conflicts_count=conflicts_count,
            benchmarks=tuple(item.benchmark_code for item in selected_benchmarks),
            from_date=from_date.isoformat() if from_date else None,
            to_date=to_date.isoformat() if to_date else None,
        )

    def _finish_error_run(
        self,
        collector_run: CollectorRun,
        started_at: datetime,
        *,
        selected_benchmarks: tuple[WorldBankBenchmark, ...],
        from_date: date | None,
        to_date: date | None,
        errors: list[str],
        raw_response_ids: list[int] | None = None,
    ) -> WorldBankPinkSheetCollectorResult:
        collector_run.finished_at = _to_utc(self.clock())
        collector_run.records_found = 0
        collector_run.records_written = 0
        collector_run.status = "error"
        collector_run.error_message = "\n".join(errors)
        return _result_from_run(
            self,
            collector_run,
            started_at,
            selected_benchmarks=selected_benchmarks,
            from_date=from_date,
            to_date=to_date,
            benchmarks_success=0,
            benchmarks_failed=len(selected_benchmarks),
            skipped_existing=0,
            conflicts_count=0,
            raw_response_ids=raw_response_ids or [],
            errors_count=len(errors),
        )

    def _collect_benchmark(
        self,
        session: Session,
        source: Source,
        collector_run: CollectorRun,
        raw_response: RawResponse,
        workbook: Any,
        benchmark: WorldBankBenchmark,
        *,
        from_date: date | None,
        to_date: date | None,
        fetched_at: datetime,
    ) -> _BenchmarkStats:
        stats = _BenchmarkStats(benchmark_code=benchmark.benchmark_code)
        try:
            rows, source_column = parse_world_bank_benchmark(
                workbook,
                benchmark=benchmark,
                from_date=from_date,
                to_date=to_date,
            )
        except WorldBankPinkSheetParseError as exc:
            _write_quality_check(
                session,
                source_id=source.id,
                table_name="commodity_benchmarks",
                check_name="commodity_benchmark_mapping_found",
                status="fail",
                severity="error",
                details={"benchmark_code": benchmark.benchmark_code, "error": str(exc)},
                checked_at=fetched_at,
            )
            stats.errors.append(f"{benchmark.benchmark_code}: {exc}")
            return stats

        _write_quality_check(
            session,
            source_id=source.id,
            table_name="commodity_benchmarks",
            check_name="commodity_benchmark_mapping_found",
            status="pass",
            severity="info",
            details={
                "benchmark_code": benchmark.benchmark_code,
                "sheet_name": benchmark.sheet_name,
                "source_column": source_column,
            },
            checked_at=fetched_at,
        )
        _write_quality_check(
            session,
            source_id=source.id,
            table_name="commodity_benchmarks",
            check_name="commodity_benchmark_parsed_rows_gt_zero",
            status="pass" if rows else "fail",
            severity="info" if rows else "error",
            details={"benchmark_code": benchmark.benchmark_code, "rows_count": len(rows)},
            checked_at=fetched_at,
        )
        if not rows:
            stats.errors.append(f"{benchmark.benchmark_code}: no benchmark rows parsed")
            return stats

        for row in rows:
            stats.records_found += 1
            source_record_hash = build_commodity_benchmark_source_record_hash(
                source_code=self.source_code,
                benchmark_code=benchmark.benchmark_code,
                frequency=benchmark.frequency,
                period_start=row.period_start,
                period_end=row.period_end,
                value=row.value,
                currency=benchmark.currency,
                unit=benchmark.unit,
            )
            _write_row_quality_checks(
                session,
                source_id=source.id,
                raw_response_id=raw_response.id,
                row=row,
                source_record_hash=source_record_hash,
                checked_at=fetched_at,
            )
            existing = _find_existing_benchmark(
                session,
                source_id=source.id,
                benchmark_code=benchmark.benchmark_code,
                frequency=benchmark.frequency,
                period_start=row.period_start,
                period_end=row.period_end,
            )
            if existing is not None:
                if existing.source_record_hash == source_record_hash:
                    stats.skipped_existing += 1
                    continue
                stats.conflicts_count += 1
                _write_quality_check(
                    session,
                    source_id=source.id,
                    table_name="commodity_benchmarks",
                    check_name="commodity_benchmark_existing_record_hash_changed",
                    status="fail",
                    severity="warning",
                    details={
                        "benchmark_code": benchmark.benchmark_code,
                        "frequency": benchmark.frequency,
                        "period_start": row.period_start.isoformat(),
                        "period_end": row.period_end.isoformat(),
                        "existing_id": existing.id,
                        "existing_value": _canonical_decimal(existing.value),
                        "incoming_value": _canonical_decimal(row.value),
                        "existing_source_record_hash": existing.source_record_hash,
                        "incoming_source_record_hash": source_record_hash,
                        "existing_raw_response_id": existing.raw_response_id,
                        "incoming_raw_response_id": raw_response.id,
                        "policy_decision": "rejected_hash_conflict_no_overwrite",
                    },
                    checked_at=fetched_at,
                )
                continue

            session.add(
                CommodityBenchmark(
                    source_id=source.id,
                    raw_response_id=raw_response.id,
                    collector_run_id=collector_run.id,
                    benchmark_code=benchmark.benchmark_code,
                    external_id=benchmark.external_id,
                    benchmark_name=benchmark.benchmark_name,
                    benchmark_category=benchmark.benchmark_category,
                    commodity_family=benchmark.commodity_family,
                    geography=benchmark.geography,
                    frequency=benchmark.frequency,
                    period_start=row.period_start,
                    period_end=row.period_end,
                    observed_at=row.observed_at,
                    published_at=None,
                    fetched_at=fetched_at,
                    value=row.value,
                    currency=benchmark.currency,
                    unit=benchmark.unit,
                    normalized_value=row.value,
                    normalized_currency=benchmark.normalized_currency,
                    normalized_unit=benchmark.normalized_unit,
                    source_record_hash=source_record_hash,
                    source_payload_hash=raw_response.sha256,
                    metadata_json=_metadata_json(row=row, raw_response=raw_response),
                )
            )
            stats.records_written += 1
            session.flush()
        return stats

    def _save_raw_response(
        self,
        session: Session,
        source: Source,
        collector_run: CollectorRun,
        response: httpx.Response,
        fetched_at: datetime,
    ) -> RawResponse:
        content_type = response.headers.get("content-type")
        raw_result = self.raw_store.save_bytes(
            source_code=self.source_code,
            collector_run_id=collector_run.id,
            payload=response.content,
            fetched_at=fetched_at,
            extension="xlsx",
            content_type=content_type,
        )
        raw_response = RawResponse(
            source_id=source.id,
            collector_run_id=collector_run.id,
            fetched_at=fetched_at,
            url=str(response.url),
            method="GET",
            status_code=response.status_code,
            content_type=content_type,
            storage_path=raw_result.storage_path,
            sha256=raw_result.sha256,
            bytes_size=raw_result.bytes_size,
            parser_version=self.parser_version,
        )
        session.add(raw_response)
        session.flush()
        return raw_response


def assess_world_bank_response(
    *,
    payload: bytes,
    content_type: str | None,
    status_code: int | None,
) -> list[BenchmarkResponseCheck]:
    return [
        BenchmarkResponseCheck(
            check_name="commodity_benchmark_raw_response_present",
            status="pass" if payload is not None else "fail",
            severity="error" if payload is None else "info",
            details={"status_code": status_code},
        ),
        BenchmarkResponseCheck(
            check_name="commodity_benchmark_raw_response_non_empty",
            status="pass" if payload else "fail",
            severity="error" if not payload else "info",
            details={"bytes_size": len(payload), "status_code": status_code},
        ),
        BenchmarkResponseCheck(
            check_name="commodity_benchmark_content_type_expected",
            status="pass" if _is_expected_content_type(content_type) else "fail",
            severity="warning" if not _is_expected_content_type(content_type) else "info",
            details={"content_type": content_type},
        ),
        BenchmarkResponseCheck(
            check_name="commodity_benchmark_response_size_within_limit",
            status="pass" if len(payload) <= MAX_RESPONSE_BYTES else "fail",
            severity="warning" if len(payload) > MAX_RESPONSE_BYTES else "info",
            details={"bytes_size": len(payload), "limit_bytes": MAX_RESPONSE_BYTES},
        ),
    ]


def parse_world_bank_benchmark(
    workbook: Any,
    *,
    benchmark: WorldBankBenchmark,
    from_date: date | None = None,
    to_date: date | None = None,
) -> tuple[list[ParsedCommodityBenchmark], str]:
    if benchmark.sheet_name not in workbook.sheetnames:
        raise WorldBankPinkSheetParseError(f"missing sheet: {benchmark.sheet_name}")
    worksheet = workbook[benchmark.sheet_name]
    source_column_index, source_column = _find_source_column(worksheet, benchmark)
    rows: list[ParsedCommodityBenchmark] = []
    for row in worksheet.iter_rows():
        if not row:
            continue
        period_start = _parse_month_cell(row[0].value)
        if period_start is None:
            continue
        if from_date is not None and period_start < from_date:
            continue
        if to_date is not None and period_start > to_date:
            continue
        if len(row) < source_column_index:
            continue
        try:
            value = _to_decimal(row[source_column_index - 1].value)
        except WorldBankPinkSheetParseError:
            continue
        period_end = date(period_start.year, period_start.month, monthrange(period_start.year, period_start.month)[1])
        rows.append(
            ParsedCommodityBenchmark(
                benchmark=benchmark,
                period_start=period_start,
                period_end=period_end,
                observed_at=datetime.combine(period_end, time.min, tzinfo=timezone.utc),
                value=value,
                source_sheet=benchmark.sheet_name,
                source_column=source_column,
                raw={
                    "source_period": row[0].value.isoformat() if isinstance(row[0].value, date) else str(row[0].value),
                    "source_value": _json_value(row[source_column_index - 1].value),
                },
            )
        )
    return rows, source_column


def build_commodity_benchmark_source_record_hash(
    *,
    source_code: str,
    benchmark_code: str,
    frequency: str,
    period_start: date,
    period_end: date,
    value: Decimal,
    currency: str | None,
    unit: str,
) -> str:
    return stable_json_hash(
        {
            "source_code": source_code,
            "benchmark_code": benchmark_code,
            "frequency": frequency,
            "period_start": period_start.isoformat(),
            "period_end": period_end.isoformat(),
            "value": _canonical_decimal(value),
            "currency": currency,
            "unit": unit,
        }
    )


def _find_source_column(worksheet: Worksheet, benchmark: WorldBankBenchmark) -> tuple[int, str]:
    candidates = {_normalize_header(item) for item in benchmark.source_column_candidates}
    for row in worksheet.iter_rows(min_row=1, max_row=12):
        for cell in row:
            if cell.value is None:
                continue
            normalized = _normalize_header(str(cell.value))
            if normalized in candidates:
                return int(cell.column), str(cell.value)
    raise WorldBankPinkSheetParseError(
        f"source column not found for {benchmark.benchmark_code}: {benchmark.source_column_candidates}"
    )


def _normalize_header(value: str) -> str:
    return " ".join(value.replace("*", "").split()).strip().lower()


def _parse_month_cell(value: Any) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return date(value.year, value.month, 1)
    if isinstance(value, date):
        return date(value.year, value.month, 1)
    text = str(value).strip()
    if len(text) == 7 and text[4].upper() == "M":
        try:
            return date(int(text[:4]), int(text[5:]), 1)
        except ValueError:
            return None
    return None


def _to_decimal(value: Any) -> Decimal:
    if value is None:
        raise WorldBankPinkSheetParseError("missing value")
    text = str(value).strip()
    if text in MISSING_VALUE_MARKERS:
        raise WorldBankPinkSheetParseError("missing value")
    try:
        return Decimal(text.replace(",", ""))
    except (InvalidOperation, ValueError) as exc:
        raise WorldBankPinkSheetParseError(f"bad numeric value: {value!r}") from exc


def _find_existing_benchmark(
    session: Session,
    *,
    source_id: int,
    benchmark_code: str,
    frequency: str,
    period_start: date,
    period_end: date,
) -> CommodityBenchmark | None:
    return session.scalar(
        select(CommodityBenchmark)
        .where(
            CommodityBenchmark.source_id == source_id,
            CommodityBenchmark.benchmark_code == benchmark_code,
            CommodityBenchmark.frequency == frequency,
            CommodityBenchmark.period_start == period_start,
            CommodityBenchmark.period_end == period_end,
        )
        .limit(1)
    )


def _write_response_checks(
    session: Session,
    *,
    source_id: int,
    raw_response_id: int,
    checks: list[BenchmarkResponseCheck],
    checked_at: datetime,
) -> None:
    for check in checks:
        details = {**check.details, "raw_response_id": raw_response_id}
        _write_quality_check(
            session,
            source_id=source_id,
            table_name="raw_responses",
            check_name=check.check_name,
            status=check.status,
            severity=check.severity,
            details=details,
            checked_at=checked_at,
        )


def _write_row_quality_checks(
    session: Session,
    *,
    source_id: int,
    raw_response_id: int,
    row: ParsedCommodityBenchmark,
    source_record_hash: str,
    checked_at: datetime,
) -> None:
    value_positive = row.value > 0
    checks = [
        ("commodity_benchmark_value_is_not_null", "pass", "info", {"value": _canonical_decimal(row.value)}),
        (
            "commodity_benchmark_value_positive",
            "pass" if value_positive else "fail",
            "info" if value_positive else "warning",
            {"value": _canonical_decimal(row.value)},
        ),
        (
            "commodity_benchmark_source_record_hash_present",
            "pass" if source_record_hash else "fail",
            "info" if source_record_hash else "error",
            {"source_record_hash_present": bool(source_record_hash)},
        ),
        (
            "commodity_benchmark_raw_linkage_present",
            "pass" if raw_response_id else "fail",
            "info" if raw_response_id else "error",
            {"raw_response_id": raw_response_id},
        ),
    ]
    for check_name, status, severity, details in checks:
        _write_quality_check(
            session,
            source_id=source_id,
            table_name="commodity_benchmarks",
            check_name=check_name,
            status=status,
            severity=severity,
            details={
                **details,
                "benchmark_code": row.benchmark.benchmark_code,
                "period_start": row.period_start.isoformat(),
                "period_end": row.period_end.isoformat(),
            },
            checked_at=checked_at,
        )


def _write_quality_check(
    session: Session,
    *,
    source_id: int | None,
    table_name: str,
    check_name: str,
    status: str,
    severity: str,
    details: dict[str, Any],
    checked_at: datetime,
) -> None:
    session.add(
        DataQualityCheck(
            source_id=source_id,
            table_name=table_name,
            check_name=check_name,
            checked_at=checked_at,
            status=status,
            severity=severity,
            details_json=details,
        )
    )
    session.flush()


def _metadata_json(*, row: ParsedCommodityBenchmark, raw_response: RawResponse) -> dict[str, Any]:
    benchmark = row.benchmark
    return {
        "source_file_url": SOURCE_FILE_URL,
        "source_document_url": SOURCE_DOCUMENT_URL,
        "source_sheet": row.source_sheet,
        "source_column": row.source_column,
        "benchmark_code": benchmark.benchmark_code,
        "parser_version": PARSER_VERSION,
        "unit_note": benchmark.unit_note,
        "frequency": benchmark.frequency,
        "period_semantics": "period_start is first calendar day; period_end/observed_at use month end",
        "filter_semantics": "from-date/to-date filter period_start",
        "raw_response_id": raw_response.id,
        "raw": row.raw,
    }


def _select_benchmarks(
    benchmarks: tuple[WorldBankBenchmark, ...],
    benchmark_code: str | None,
) -> tuple[WorldBankBenchmark, ...]:
    if benchmark_code is None:
        return benchmarks
    if benchmark_code not in WORLD_BANK_BENCHMARKS_BY_CODE:
        raise ValueError(f"unknown World Bank Pink Sheet benchmark: {benchmark_code}")
    selected = tuple(item for item in benchmarks if item.benchmark_code == benchmark_code)
    if not selected:
        raise ValueError(f"benchmark is not enabled for this collector: {benchmark_code}")
    return selected


def _normalize_run_type(run_type: str) -> str:
    normalized = run_type.strip().lower()
    if normalized not in {"manual", "scheduled", "backfill"}:
        raise ValueError("run_type must be manual, scheduled, or backfill")
    return normalized


def _result_from_run(
    collector: WorldBankPinkSheetCollector,
    collector_run: CollectorRun,
    started_at: datetime,
    *,
    selected_benchmarks: tuple[WorldBankBenchmark, ...],
    from_date: date | None,
    to_date: date | None,
    benchmarks_success: int,
    benchmarks_failed: int,
    skipped_existing: int,
    conflicts_count: int,
    raw_response_ids: list[int],
    errors_count: int,
) -> WorldBankPinkSheetCollectorResult:
    return WorldBankPinkSheetCollectorResult(
        run_id=collector_run.id,
        source_code=collector.source_code,
        collector_name=collector.collector_name,
        parser_version=collector.parser_version,
        started_at=started_at,
        finished_at=collector_run.finished_at or started_at,
        status=collector_run.status,
        records_found=collector_run.records_found,
        records_written=collector_run.records_written,
        raw_response_ids=raw_response_ids,
        error_message=collector_run.error_message,
        errors_count=errors_count,
        benchmarks_requested=len(selected_benchmarks),
        benchmarks_success=benchmarks_success,
        benchmarks_failed=benchmarks_failed,
        skipped_existing=skipped_existing,
        conflicts_count=conflicts_count,
        benchmarks=tuple(item.benchmark_code for item in selected_benchmarks),
        from_date=from_date.isoformat() if from_date else None,
        to_date=to_date.isoformat() if to_date else None,
    )


def _is_expected_content_type(content_type: str | None) -> bool:
    if not content_type:
        return False
    normalized = content_type.split(";")[0].strip().lower()
    return normalized in {
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel",
        "application/octet-stream",
    }


def _to_utc(value: datetime) -> datetime:
    if value.tzinfo is None:
        return value.replace(tzinfo=timezone.utc)
    return value.astimezone(timezone.utc)


def _canonical_decimal(value: Decimal | None) -> str | None:
    if value is None:
        return None
    return format(value.normalize(), "f")


def _json_value(value: Any) -> str | int | float | None:
    if value is None:
        return None
    if isinstance(value, Decimal):
        return _canonical_decimal(value)
    if isinstance(value, datetime | date):
        return value.isoformat()
    if isinstance(value, int | float | str):
        return value
    return str(value)
